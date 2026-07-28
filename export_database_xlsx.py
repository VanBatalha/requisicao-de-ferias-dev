from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from sqlalchemy import create_engine, inspect, text

from ferias_app.config import get_settings


INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _schema_name() -> str:
    schema = (os.getenv("DB_SCHEMA") or "app_ferias").strip()
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", schema):
        raise ValueError(f"DB_SCHEMA invalido: {schema!r}")
    return schema


def _quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _safe_sheet_name(table_name: str, used: set[str]) -> str:
    base = INVALID_SHEET_CHARS.sub("_", table_name).strip() or "Tabela"
    base = base[:31]
    candidate = base
    suffix = 1
    while candidate.lower() in used:
        tail = f"_{suffix}"
        candidate = (base[: 31 - len(tail)] + tail)[:31]
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, memoryview):
        return bytes(value).hex()
    return value


def _table_order_key(name: str) -> tuple[int, str]:
    lowered = name.lower()
    is_backup = lowered.startswith("z_backup") or lowered.startswith("z_")
    return (1 if is_backup else 0, lowered)


def _ordered_columns(inspector, schema: str, table: str) -> tuple[list[str], list[str]]:
    columns = [column["name"] for column in inspector.get_columns(table, schema=schema)]
    pk = inspector.get_pk_constraint(table, schema=schema).get("constrained_columns") or []
    if "id" in columns:
        order = ["id"]
    elif pk:
        order = [column for column in pk if column in columns]
    else:
        order = []
    return columns, order


def _iter_rows(connection, sql: str, batch_size: int = 1000) -> Iterable[tuple[Any, ...]]:
    result = connection.execution_options(stream_results=True).execute(text(sql))
    while True:
        rows = result.fetchmany(batch_size)
        if not rows:
            break
        for row in rows:
            yield tuple(row)


def export_database(output_path: Path) -> Path:
    settings = get_settings()
    if not settings.database_url:
        raise RuntimeError("Banco nao configurado. Defina DATABASE_URL ou as variaveis PG_*.")

    schema = _schema_name()
    engine = create_engine(
        settings.database_url,
        pool_pre_ping=True,
        connect_args={"connect_timeout": 10, "application_name": "ferias_app_xlsx_export"},
    )
    inspector = inspect(engine)
    tables = sorted(inspector.get_table_names(schema=schema), key=_table_order_key)
    if not tables:
        raise RuntimeError(f"Nenhuma tabela encontrada no schema {schema!r}.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    used_sheet_names: set[str] = set()
    index_rows: list[tuple[str, str, int, str]] = []

    with engine.connect() as connection:
        connection.execute(text(f"SET search_path TO {_quote(schema)}, public"))
        for table in tables:
            columns, order_columns = _ordered_columns(inspector, schema, table)
            sheet_name = _safe_sheet_name(table, used_sheet_names)
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.freeze_panes = "A2"

            header_cells = []
            for column in columns:
                cell = WriteOnlyCell(worksheet, value=column)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                header_cells.append(cell)
            worksheet.append(header_cells)

            select_columns = ", ".join(_quote(column) for column in columns)
            sql = f"SELECT {select_columns} FROM {_quote(schema)}.{_quote(table)}"
            if order_columns:
                sql += " ORDER BY " + ", ".join(_quote(column) + " ASC NULLS LAST" for column in order_columns)

            row_count = 0
            for row in _iter_rows(connection, sql):
                worksheet.append([_excel_value(value) for value in row])
                row_count += 1

            order_description = ", ".join(order_columns) if order_columns else "sem chave de ordenacao"
            index_rows.append((table, sheet_name, row_count, order_description))

    index_sheet = workbook.create_sheet(_safe_sheet_name("INDICE", used_sheet_names))
    index_sheet.append(["tabela", "aba", "linhas", "ordenacao"])
    for table, sheet_name, row_count, order_description in index_rows:
        index_sheet.append([table, sheet_name, row_count, order_description])

    workbook.save(output_path)
    engine.dispose()
    return output_path


def main() -> None:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    parser = argparse.ArgumentParser(
        description="Exporta todas as tabelas do PostgreSQL para XLSX, ordenando por ID ou chave primaria."
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=f"export_app_ferias_{stamp}.xlsx",
        help="Caminho do arquivo XLSX de saida.",
    )
    args = parser.parse_args()
    path = export_database(Path(args.output).expanduser().resolve())
    print(path)


if __name__ == "__main__":
    main()
