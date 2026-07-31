from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _text(item: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = item.get(key)
        if value not in (None, ""):
            return str(value)
    return ""


def _num(item: dict[str, Any], key: str) -> float:
    try:
        return float(item.get(key) or 0)
    except (TypeError, ValueError):
        return 0.0


def _auto_width(ws, min_width: int = 10, max_width: int = 48) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        width = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, min_width), max_width)


def criar_relatorio_saldos_xlsx(
    colaboradores: Iterable[dict[str, Any]],
    *,
    status_filtro: str = "",
    busca: str = "",
) -> bytes:
    """Gera relatório de saldos com resumo executivo e detalhamento formatado."""
    rows = list(colaboradores)
    wb = Workbook()
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws = wb.create_sheet("Saldos")

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_secao = PatternFill("solid", fgColor="D9EAF7")
    fill_cabecalho = PatternFill("solid", fgColor="5B9BD5")
    fill_grupo_regular = PatternFill("solid", fgColor="DDEBF7")
    fill_grupo_premium = PatternFill("solid", fgColor="E4DFEC")
    fill_total = PatternFill("solid", fgColor="E2F0D9")
    font_branca = Font(color="FFFFFF", bold=True)
    font_titulo = Font(color="FFFFFF", bold=True, size=14)
    font_negrito = Font(bold=True)
    borda = Border(bottom=Side(style="thin", color="D9E1F2"))
    borda_cabecalho = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    # Resumo
    ws_resumo.merge_cells("A1:E1")
    ws_resumo["A1"] = "Relatório de Saldos dos Colaboradores"
    ws_resumo["A1"].fill = fill_titulo
    ws_resumo["A1"].font = font_titulo
    ws_resumo["A1"].alignment = Alignment(horizontal="center")
    ws_resumo.row_dimensions[1].height = 24

    meta = [
        ("Status filtrado", status_filtro or "Todos"),
        ("Pesquisa", busca or "Sem filtro"),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for idx, (label, value) in enumerate(meta, start=3):
        ws_resumo.cell(idx, 1, label).font = font_negrito
        ws_resumo.cell(idx, 2, value)

    ws_resumo["A8"] = "Indicadores"
    ws_resumo["A8"].fill = fill_secao
    ws_resumo["A8"].font = font_negrito
    ws_resumo.merge_cells("A8:B8")

    indicadores = [
        ("Total de colaboradores", "=MAX(COUNTA(Saldos!A:A)-4,0)"),
        ("Colaboradores ativos", '=COUNTIF(Saldos!H:H,"ATIVO")'),
        ("Colaboradores inativos", '=COUNTIF(Saldos!H:H,"INATIVO")'),
        ("Saldo regular disponível", "=SUM(Saldos!L:L)"),
        ("Saldo Certariana disponível", "=SUM(Saldos!P:P)"),
        ("Saldo total disponível", "=SUM(Saldos!Q:Q)"),
    ]
    for idx, (label, value) in enumerate(indicadores, start=9):
        ws_resumo.cell(idx, 1, label)
        ws_resumo.cell(idx, 2, value)
        ws_resumo.cell(idx, 1).border = borda
        ws_resumo.cell(idx, 2).border = borda
        if idx >= 12:
            ws_resumo.cell(idx, 2).number_format = '0.00'

    resumo_header_row = 18
    resumo_headers = ["Matrícula", "Colaborador", "Status", "Saldo regular", "Saldo Certariana", "Saldo total"]
    for col_idx, header in enumerate(resumo_headers, start=1):
        cell = ws_resumo.cell(resumo_header_row, col_idx, header)
        cell.fill = fill_cabecalho
        cell.font = font_branca
        cell.alignment = Alignment(horizontal="center")

    rows_sorted = sorted(rows, key=lambda item: (
        _text(item, "nome_completo", "NOME COMPLETO", "nome").casefold(),
        _text(item, "matricula", "MATRICULA", "MATRÍCULA"),
    ))
    for row_idx, item in enumerate(rows_sorted, start=resumo_header_row + 1):
        vals = [
            _text(item, "matricula", "MATRICULA", "MATRÍCULA"),
            _text(item, "nome_completo", "NOME COMPLETO", "nome"),
            _text(item, "status", "STATUS"),
            _num(item, "saldo_regular"),
            _num(item, "saldo_premium"),
            _num(item, "saldo_total"),
        ]
        for col_idx, value in enumerate(vals, start=1):
            ws_resumo.cell(row_idx, col_idx, value)
        for col_idx in range(4, 7):
            ws_resumo.cell(row_idx, col_idx).number_format = '0.00'

    # Detalhes
    ws.merge_cells("A1:Q1")
    ws["A1"] = "Saldos detalhados por colaborador"
    ws["A1"].fill = fill_titulo
    ws["A1"].font = font_titulo
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Filtros: status {status_filtro or 'Todos'} | pesquisa {busca or 'Sem filtro'}"
    ws["A2"].alignment = Alignment(horizontal="left")
    ws["A2"].font = Font(italic=True, color="666666")

    ws.merge_cells("I3:L3")
    ws["I3"] = "Férias regulares"
    ws["I3"].fill = fill_grupo_regular
    ws["I3"].font = font_negrito
    ws["I3"].alignment = Alignment(horizontal="center")
    ws.merge_cells("M3:P3")
    ws["M3"] = "Licença Certariana"
    ws["M3"].fill = fill_grupo_premium
    ws["M3"].font = font_negrito
    ws["M3"].alignment = Alignment(horizontal="center")

    headers = [
        "Matrícula", "Nome", "E-mail", "Cargo", "Setor", "Unidade", "Empresa", "Status",
        "Direito", "Utilizado", "Reservado", "Disponível",
        "Direito", "Utilizado", "Reservado", "Disponível",
        "Saldo total",
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.fill = fill_cabecalho
        cell.font = font_branca
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda_cabecalho

    data_start = header_row + 1
    for row_idx, item in enumerate(rows_sorted, start=data_start):
        values = [
            _text(item, "matricula", "MATRICULA", "MATRÍCULA"),
            _text(item, "nome_completo", "NOME COMPLETO", "nome"),
            _text(item, "email", "EMAIL DA EMPRESA"),
            _text(item, "cargo", "CARGO"),
            _text(item, "setor", "SETOR"),
            _text(item, "unidade", "UNIDADE"),
            _text(item, "empresa", "EMPRESA"),
            _text(item, "status", "STATUS"),
            _num(item, "saldo_regular_direito"),
            _num(item, "saldo_regular_usado"),
            _num(item, "saldo_regular_reservado"),
            _num(item, "saldo_regular"),
            _num(item, "saldo_premium_direito"),
            _num(item, "saldo_premium_usado"),
            _num(item, "saldo_premium_reservado"),
            _num(item, "saldo_premium"),
            _num(item, "saldo_total"),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
        for col_idx in range(9, 18):
            ws.cell(row_idx, col_idx).number_format = '0.00'

    total_row = data_start + len(rows_sorted) + 1
    ws.cell(total_row, 1, "TOTAL").font = font_negrito
    ws.cell(total_row, 1).fill = fill_total
    for col_idx in range(9, 18):
        letter = get_column_letter(col_idx)
        end_data = max(data_start, total_row - 2)
        cell = ws.cell(total_row, col_idx, f"=SUM({letter}{data_start}:{letter}{end_data})")
        cell.font = font_negrito
        cell.fill = fill_total
        cell.number_format = '0.00'

    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = f"A{header_row}:Q{max(header_row, data_start + len(rows_sorted) - 1)}"
    ws_resumo.freeze_panes = f"A{resumo_header_row + 1}"
    ws_resumo.auto_filter.ref = f"A{resumo_header_row}:F{max(resumo_header_row, resumo_header_row + len(rows_sorted))}"
    ws.sheet_view.showGridLines = False
    ws_resumo.sheet_view.showGridLines = False
    ws.row_dimensions[header_row].height = 34
    _auto_width(ws_resumo, max_width=42)
    _auto_width(ws, max_width=45)
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 34

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
