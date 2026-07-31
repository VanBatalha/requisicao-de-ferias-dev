"""Relatório de solicitações gerado diretamente pelo PostgreSQL.

V63:
- conexão curta e exclusiva, sem disputar o pool SQLAlchemy das telas;
- perfil e matrícula lidos da sessão, sem Smartsheet e sem busca por e-mail;
- limite do PostgreSQL e limite rígido no processo para impedir timeout do Gunicorn;
- cache local do último resultado por usuário/período como contingência;
- DP/ADMIN consultam todos; gestores consultam somente sua equipe.
"""
from __future__ import annotations

from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal
import hashlib
from io import BytesIO
import json
import os
from pathlib import Path
import re
import signal
import threading
import time
import uuid
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import psycopg2
from psycopg2 import InterfaceError, OperationalError
from psycopg2.extras import RealDictCursor

from ..config import get_settings
from ..logging_config import get_logger

log = get_logger(__name__)

_CACHE_DIR = Path(os.getenv("RELATORIO_CACHE_DIR") or "/tmp/ferias_app_relatorios")
_CACHE_TTL_SECONDS = max(60, int(os.getenv("RELATORIO_CACHE_TTL_SECONDS") or "600"))
_QUERY_HARD_TIMEOUT_SECONDS = max(3, int(os.getenv("RELATORIO_QUERY_TIMEOUT_SECONDS") or "9"))


class RelatorioAcessoNegado(PermissionError):
    """Usuário não possui perfil DP/ADMIN nem equipe cadastrada."""


class RelatorioTempoExcedido(TimeoutError):
    """A consulta ultrapassou o limite seguro do request web."""


def _schema_name() -> str:
    schema = (os.getenv("DB_SCHEMA") or "app_ferias").strip()
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", schema):
        return "app_ferias"
    return schema


def _quote_ident(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _numero(valor: Any) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, Decimal):
        return float(valor)
    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def _status_key(status: Any) -> tuple[str, str]:
    original = str(status or "PENDENTE").strip().upper()
    normalizado = (
        original.replace("Á", "A").replace("Ã", "A").replace("Â", "A")
        .replace("É", "E").replace("Ê", "E").replace("Í", "I")
        .replace("Ó", "O").replace("Ô", "O").replace("Õ", "O")
        .replace("Ú", "U").replace("Ç", "C")
    )
    if "APROV" in normalizado:
        return original, "aprovada"
    if any(valor in normalizado for valor in ("REPROV", "REJEIT", "NEGAD", "CANCEL")):
        return original, "negada"
    if "ANALISE" in normalizado:
        return original, "em_analise"
    if "RESERV" in normalizado:
        return original, "reservada"
    return original, "pendente"


def _normalizar_perfil(value: Any) -> str:
    value = str(value or "").strip().upper()
    if value in {"ADMIN", "ADMINISTRADOR"}:
        return "ADMIN"
    if value in {"DP", "RH"}:
        return "DP"
    return "USER"


def _normalizar_escopo(value: Any, perfil: str) -> str:
    raw = str(value or "").strip().lower().replace("-", "_")
    aliases = {
        "usuario": "colaborador",
        "colaborador_selecionado": "colaborador",
        "minha_equipe": "equipe",
        "equipe_do_gestor": "equipe",
        "gestor": "equipe_gestor",
        "equipe_selecionada": "equipe_gestor",
        "all": "todos",
    }
    raw = aliases.get(raw, raw)
    if perfil in {"ADMIN", "DP"}:
        return raw if raw in {"todos", "colaborador", "equipe_gestor"} else "todos"
    return "colaborador" if raw == "colaborador" else "equipe"


def _resultado_vazio(
    *, request_id: str, mes: Optional[int], ano: Optional[int], escopo: str,
    referencia: str, total_colaboradores_escopo: Optional[int],
    escopo_label: str = "", colaborador_referencia: str = "",
) -> Dict[str, Any]:
    return {
        "ok": True,
        "request_id": request_id,
        "total": 0,
        "por_colaborador": {},
        "resumo_status": {"pendente": 0, "em_analise": 0, "aprovada": 0, "negada": 0, "reservada": 0},
        "total_dias": 0,
        "mes": mes,
        "ano": ano,
        "escopo": escopo,
        "escopo_label": escopo_label,
        "gestor_referencia": referencia,
        "colaborador_referencia": colaborador_referencia,
        "total_colaboradores_escopo": total_colaboradores_escopo,
        "total_colaboradores_com_lancamento": 0,
        "from_cache": False,
    }


def _abrir_conexao():
    settings = get_settings()
    if not settings.database_url:
        raise ValueError("PostgreSQL não configurado para gerar o relatório.")
    return psycopg2.connect(
        settings.database_url,
        connect_timeout=4,
        application_name="ferias_app_relatorio_v63",
        options=(
            "-c statement_timeout=6000 "
            "-c lock_timeout=1500 "
            "-c idle_in_transaction_session_timeout=7000"
        ),
    )


@contextmanager
def _limite_rigido(segundos: int, etapa: str):
    """Interrompe chamada bloqueada antes do timeout padrão do Gunicorn.

    Gunicorn síncrono executa o request na thread principal, onde SIGALRM é
    suportado. Em ambientes sem SIGALRM, o statement_timeout do PostgreSQL
    continua sendo a proteção disponível.
    """
    pode_usar = (
        hasattr(signal, "SIGALRM")
        and hasattr(signal, "setitimer")
        and threading.current_thread() is threading.main_thread()
    )
    if not pode_usar:
        yield
        return

    def _handler(_signum, _frame):
        raise RelatorioTempoExcedido(f"Tempo excedido em {etapa}.")

    antigo_handler = signal.getsignal(signal.SIGALRM)
    signal.signal(signal.SIGALRM, _handler)
    antigo_timer = signal.setitimer(signal.ITIMER_REAL, float(segundos))
    try:
        yield
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, antigo_handler)
        if antigo_timer and antigo_timer[0] > 0:
            signal.setitimer(signal.ITIMER_REAL, antigo_timer[0], antigo_timer[1])


def _executar(cursor, sql: str, params: list[Any] | tuple[Any, ...], etapa: str) -> None:
    with _limite_rigido(_QUERY_HARD_TIMEOUT_SECONDS, etapa):
        cursor.execute(sql, params)


def _filtro_periodo(mes: Optional[int], ano: Optional[int]) -> tuple[str, list[Any]]:
    if ano and mes:
        data_de = date(int(ano), int(mes), 1)
        data_ate = date(int(ano) + 1, 1, 1) if int(mes) == 12 else date(int(ano), int(mes) + 1, 1)
        return " AND s.data_inicio >= %s AND s.data_inicio < %s", [data_de, data_ate]
    if ano:
        return " AND s.data_inicio >= %s AND s.data_inicio < %s", [date(int(ano), 1, 1), date(int(ano) + 1, 1, 1)]
    if mes:
        return " AND EXTRACT(MONTH FROM s.data_inicio) = %s", [int(mes)]
    return "", []


def _cache_path(
    matricula: str, perfil: str, mes: Optional[int], ano: Optional[int],
    escopo: str = "", colaborador_matricula: str = "", gestor_matricula: str = "",
) -> Path:
    chave = (
        f"{matricula}|{perfil}|{mes or 0}|{ano or 0}|{escopo}|"
        f"{colaborador_matricula or ''}|{gestor_matricula or ''}"
    ).encode("utf-8")
    return _CACHE_DIR / f"{hashlib.sha256(chave).hexdigest()}.json"


def _gravar_cache(path: Path, payload: Dict[str, Any]) -> None:
    try:
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        temp = path.with_suffix(f".{os.getpid()}.tmp")
        temp.write_text(
            json.dumps({"saved_at": time.time(), "payload": payload}, ensure_ascii=False),
            encoding="utf-8",
        )
        os.replace(temp, path)
    except Exception as exc:  # cache nunca deve impedir o relatório
        log.warning("Falha ao gravar cache de relatório: %s", exc)


def _ler_cache(path: Path) -> Optional[Dict[str, Any]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        idade = time.time() - float(data.get("saved_at") or 0)
        if idade < 0 or idade > _CACHE_TTL_SECONDS:
            return None
        payload = data.get("payload")
        if not isinstance(payload, dict) or not payload.get("ok"):
            return None
        payload = dict(payload)
        payload["from_cache"] = True
        payload["cache_age_seconds"] = int(idade)
        payload["warning"] = (
            "O PostgreSQL demorou a responder. Foi exibido o último relatório "
            "gerado para este mesmo usuário e período."
        )
        return payload
    except Exception:
        return None


def _usar_cache_ou_erro(
    *, cache_path: Path, request_id: str, etapa: str, exc: Exception,
    inicio_total: float,
) -> Dict[str, Any]:
    log.exception("RELATORIO[%s] falhou em '%s' após %.3fs", request_id, etapa, time.monotonic() - inicio_total)
    cached = _ler_cache(cache_path)
    if cached:
        cached["request_id"] = request_id
        log.warning("RELATORIO[%s] respondido pelo cache de contingência", request_id)
        return cached
    return {
        "ok": False,
        "request_id": request_id,
        "message": (
            "Não foi possível obter resposta do PostgreSQL a tempo. "
            f"Etapa: {etapa}. Código: {request_id}."
        ),
        "detail": str(exc),
    }


def gerar_relatorio_lancamento(
    usuario_matricula: str,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    perfil_sessao: Optional[str] = None,
    escopo_solicitado: Optional[str] = None,
    colaborador_matricula: Optional[str] = None,
    gestor_matricula: Optional[str] = None,
) -> Dict[str, Any]:
    """Gera relatório por matrícula, com escopo explícito e validação no backend.

    Regras:
    - DP/ADMIN: todos, colaborador selecionado ou equipe de um gestor;
    - gestor comum: sua equipe (incluindo o próprio gestor) ou um colaborador
      pertencente ao seu escopo.
    """
    request_id = uuid.uuid4().hex[:10]
    inicio_total = time.monotonic()
    matricula_usuario = str(usuario_matricula or "").strip().upper()
    perfil = _normalizar_perfil(perfil_sessao)
    escopo = _normalizar_escopo(escopo_solicitado, perfil)
    alvo_colaborador = str(colaborador_matricula or "").strip().upper()
    alvo_gestor = str(gestor_matricula or "").strip().upper()
    etapa = "abrindo conexão exclusiva"

    if not matricula_usuario:
        raise ValueError("Matrícula do usuário não foi identificada na sessão. Saia e entre novamente no sistema.")
    if escopo == "colaborador" and not alvo_colaborador:
        raise ValueError("Selecione um colaborador para gerar este relatório.")
    if escopo == "equipe_gestor" and not alvo_gestor:
        raise ValueError("Selecione um gestor para gerar o relatório da equipe.")

    schema_sql = _quote_ident(_schema_name())
    cache_path = _cache_path(
        matricula_usuario, perfil, mes, ano, escopo,
        colaborador_matricula=alvo_colaborador,
        gestor_matricula=alvo_gestor,
    )
    conn = None
    try:
        with _limite_rigido(6, etapa):
            conn = _abrir_conexao()
        conn.autocommit = True
        log.info("RELATORIO[%s] conexão exclusiva em %.3fs", request_id, time.monotonic() - inicio_total)

        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            is_admin_dp = perfil in {"ADMIN", "DP"}
            total_colaboradores_escopo: Optional[int] = None
            escopo_sql = ""
            escopo_params: list[Any] = []
            referencia = ""
            colaborador_referencia = ""

            if escopo == "todos":
                if not is_admin_dp:
                    raise RelatorioAcessoNegado("Acesso negado ao relatório de todos os colaboradores.")
                escopo_label = "Todos os colaboradores"

            elif escopo == "colaborador":
                colaborador_referencia = alvo_colaborador
                referencia = alvo_colaborador
                escopo_label = "Colaborador selecionado"
                if not is_admin_dp:
                    etapa = "validando colaborador no escopo"
                    _executar(
                        cursor,
                        f"""
                        SELECT (
                            %s = %s OR EXISTS (
                                SELECT 1
                                  FROM {schema_sql}.hierarquia_gestao h
                                 WHERE h.colaborador_matricula = %s
                                   AND (h.gestor_direto_matricula = %s OR h.gestor_superior_matricula = %s)
                            )
                        ) AS permitido
                        """,
                        (alvo_colaborador, matricula_usuario, alvo_colaborador, matricula_usuario, matricula_usuario),
                        etapa,
                    )
                    permitido = bool((cursor.fetchone() or {}).get("permitido"))
                    if not permitido:
                        raise RelatorioAcessoNegado("O colaborador selecionado não pertence à sua equipe.")
                else:
                    etapa = "validando colaborador"
                    _executar(
                        cursor,
                        f"SELECT EXISTS(SELECT 1 FROM {schema_sql}.colaboradores WHERE matricula = %s) AS existe",
                        (alvo_colaborador,),
                        etapa,
                    )
                    if not bool((cursor.fetchone() or {}).get("existe")):
                        raise ValueError("Colaborador selecionado não encontrado.")
                total_colaboradores_escopo = 1
                escopo_sql = " AND s.colaborador_matricula = %s"
                escopo_params.append(alvo_colaborador)

            else:
                gestor_ref = alvo_gestor if escopo == "equipe_gestor" else matricula_usuario
                if escopo == "equipe_gestor" and not is_admin_dp:
                    raise RelatorioAcessoNegado("Apenas DP/ADMIN podem consultar a equipe de outro gestor.")
                referencia = gestor_ref
                escopo_label = "Equipe do gestor (incluindo o gestor)"

                etapa = "validando gestor"
                _executar(
                    cursor,
                    f"SELECT EXISTS(SELECT 1 FROM {schema_sql}.colaboradores WHERE matricula = %s) AS existe",
                    (gestor_ref,),
                    etapa,
                )
                if not bool((cursor.fetchone() or {}).get("existe")):
                    raise ValueError("Gestor selecionado não encontrado.")

                etapa = "consultando hierarquia"
                marco = time.monotonic()
                _executar(
                    cursor,
                    f"""
                    SELECT count(*) AS total
                      FROM (
                            SELECT %s::text AS matricula
                            UNION
                            SELECT h.colaborador_matricula
                              FROM {schema_sql}.hierarquia_gestao h
                             WHERE h.gestor_direto_matricula = %s
                                OR h.gestor_superior_matricula = %s
                      ) equipe
                     WHERE matricula IS NOT NULL AND btrim(matricula) <> ''
                    """,
                    (gestor_ref, gestor_ref, gestor_ref),
                    etapa,
                )
                total_colaboradores_escopo = int((cursor.fetchone() or {}).get("total") or 0)
                log.info(
                    "RELATORIO[%s] hierarquia em %.3fs (%d matrícula(s), gestor incluído)",
                    request_id, time.monotonic() - marco, total_colaboradores_escopo,
                )
                escopo_sql = f"""
                    AND (
                        s.colaborador_matricula = %s
                        OR EXISTS (
                            SELECT 1
                              FROM {schema_sql}.hierarquia_gestao h
                             WHERE h.colaborador_matricula = s.colaborador_matricula
                               AND (h.gestor_direto_matricula = %s OR h.gestor_superior_matricula = %s)
                        )
                    )
                """
                escopo_params.extend([gestor_ref, gestor_ref, gestor_ref])

            etapa = "consultando solicitações"
            marco = time.monotonic()
            periodo_sql, periodo_params = _filtro_periodo(mes, ano)
            params: list[Any] = list(escopo_params) + list(periodo_params)

            _executar(
                cursor,
                f"""
                SELECT
                    s.id,
                    s.colaborador_matricula,
                    c.nome_completo,
                    s.data_inicio,
                    s.data_fim,
                    COALESCE(s.dias_solicitados, s.dias, 0) AS dias,
                    s.status,
                    COALESCE(s.tipo_solicitacao, s.solicitacao, '') AS solicitacao,
                    COALESCE(s.tipo_ferias, s.saldo_tipo, 'REGULAR') AS saldo_tipo,
                    COALESCE(s.observacoes, '') AS observacoes
                  FROM {schema_sql}.solicitacoes_ferias s
                  LEFT JOIN {schema_sql}.colaboradores c ON c.matricula = s.colaborador_matricula
                 WHERE COALESCE(s.is_ajuste, FALSE) = FALSE
                       {escopo_sql}
                       {periodo_sql}
                 ORDER BY s.data_inicio DESC NULLS LAST, s.id DESC
                """,
                params,
                etapa,
            )
            rows = cursor.fetchall()
            log.info("RELATORIO[%s] solicitações em %.3fs (%d linha(s))", request_id, time.monotonic() - marco, len(rows))

        if not rows:
            resultado = _resultado_vazio(
                request_id=request_id, mes=mes, ano=ano, escopo=escopo,
                referencia=referencia, total_colaboradores_escopo=total_colaboradores_escopo,
                escopo_label=escopo_label, colaborador_referencia=colaborador_referencia,
            )
            _gravar_cache(cache_path, resultado)
            log.info("RELATORIO[%s] concluído vazio em %.3fs", request_id, time.monotonic() - inicio_total)
            return resultado

        etapa = "montando resposta"
        por_colaborador: Dict[str, Any] = {}
        resumo_status = {"pendente": 0, "em_analise": 0, "aprovada": 0, "negada": 0, "reservada": 0}
        total_dias = 0.0
        matriculas_com_lancamento: set[str] = set()

        for row in rows:
            status_original, status_chave = _status_key(row.get("status"))
            resumo_status[status_chave] += 1
            matricula = str(row.get("colaborador_matricula") or "SEM MATRICULA").strip().upper()
            matriculas_com_lancamento.add(matricula)
            nome = str(row.get("nome_completo") or "").strip()
            chave = f"{nome} ({matricula})" if nome else matricula
            info = por_colaborador.setdefault(
                chave,
                {"matricula": matricula, "nome": nome, "solicitacoes": [], "total_dias": 0.0, "total_aprovadas": 0.0},
            )
            dias = _numero(row.get("dias"))
            data_inicio = row.get("data_inicio")
            data_fim = row.get("data_fim")
            info["solicitacoes"].append({
                "id": row.get("id"),
                "inicio": data_inicio.strftime("%d/%m/%Y") if data_inicio else "",
                "fim": data_fim.strftime("%d/%m/%Y") if data_fim else "",
                "dias": dias,
                "status": status_original,
                "solicitacao": row.get("solicitacao") or "",
                "saldo_tipo": row.get("saldo_tipo") or "REGULAR",
                "observacoes": row.get("observacoes") or "",
            })
            info["total_dias"] += dias
            if status_chave == "aprovada":
                info["total_aprovadas"] += dias
            total_dias += dias

        resultado = {
            "ok": True,
            "request_id": request_id,
            "total": len(rows),
            "por_colaborador": por_colaborador,
            "resumo_status": resumo_status,
            "total_dias": round(total_dias, 2),
            "mes": mes,
            "ano": ano,
            "escopo": escopo,
            "escopo_label": escopo_label,
            "gestor_referencia": referencia,
            "colaborador_referencia": colaborador_referencia,
            "total_colaboradores_escopo": total_colaboradores_escopo,
            "total_colaboradores_com_lancamento": len(matriculas_com_lancamento),
            "from_cache": False,
        }
        _gravar_cache(cache_path, resultado)
        log.info("RELATORIO[%s] concluído em %.3fs", request_id, time.monotonic() - inicio_total)
        return resultado

    except RelatorioAcessoNegado:
        raise
    except (OperationalError, InterfaceError, RelatorioTempoExcedido) as exc:
        return _usar_cache_ou_erro(
            cache_path=cache_path, request_id=request_id, etapa=etapa,
            exc=exc, inicio_total=inicio_total,
        )
    except Exception as exc:
        return _usar_cache_ou_erro(
            cache_path=cache_path, request_id=request_id, etapa=etapa,
            exc=exc, inicio_total=inicio_total,
        )
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


_MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}


def obter_relatorio_para_exportacao(
    usuario_matricula: str,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    perfil_sessao: Optional[str] = None,
    escopo_solicitado: Optional[str] = None,
    colaborador_matricula: Optional[str] = None,
    gestor_matricula: Optional[str] = None,
) -> Dict[str, Any]:
    """Obtém o mesmo relatório da tela, priorizando o cache já gerado."""
    matricula = str(usuario_matricula or "").strip().upper()
    perfil = _normalizar_perfil(perfil_sessao)
    escopo = _normalizar_escopo(escopo_solicitado, perfil)
    colab = str(colaborador_matricula or "").strip().upper()
    gestor = str(gestor_matricula or "").strip().upper()
    if not matricula:
        raise ValueError("Matrícula do usuário não foi identificada na sessão.")

    cached = _ler_cache(_cache_path(matricula, perfil, mes, ano, escopo, colab, gestor))
    if cached:
        cached["warning"] = "Relatório exportado a partir do resultado exibido na tela."
        return cached
    return gerar_relatorio_lancamento(
        matricula, mes, ano, perfil_sessao=perfil,
        escopo_solicitado=escopo,
        colaborador_matricula=colab,
        gestor_matricula=gestor,
    )


def _parse_data_br(value: Any):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%d/%m/%Y").date()
    except ValueError:
        return text


def _auto_width(ws, min_width: int = 10, max_width: int = 48) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        maior = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            maior = max(maior, len(str(value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(maior + 2, min_width), max_width)


def criar_relatorio_lancamento_xlsx(relatorio: Dict[str, Any]) -> bytes:
    """Cria o arquivo XLSX do relatório visualizado na tela."""
    if not relatorio or not relatorio.get("ok"):
        raise ValueError((relatorio or {}).get("message") or "Relatório indisponível para exportação.")

    wb = Workbook()
    # Garante que Excel/LibreOffice recalculem os indicadores ao abrir.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_detalhes = wb.create_sheet("Solicitações")

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_secao = PatternFill("solid", fgColor="D9EAF7")
    fill_cabecalho = PatternFill("solid", fgColor="5B9BD5")
    font_branca = Font(color="FFFFFF", bold=True)
    font_titulo = Font(color="FFFFFF", bold=True, size=14)
    font_negrito = Font(bold=True)
    borda_fina = Border(bottom=Side(style="thin", color="D9E1F2"))

    ws_resumo.merge_cells("A1:D1")
    ws_resumo["A1"] = "Relatório de Solicitações de Férias"
    ws_resumo["A1"].fill = fill_titulo
    ws_resumo["A1"].font = font_titulo
    ws_resumo["A1"].alignment = Alignment(horizontal="center")
    ws_resumo.row_dimensions[1].height = 24

    mes = relatorio.get("mes")
    ano = relatorio.get("ano")
    periodo = "Todos os períodos"
    if mes and ano:
        periodo = f"{_MESES_PT.get(int(mes), mes)} de {ano}"
    elif ano:
        periodo = f"Ano de {ano}"
    elif mes:
        periodo = f"Mês de {_MESES_PT.get(int(mes), mes)} (todos os anos)"

    metadados = [
        ("Período", periodo),
        ("Escopo", relatorio.get("escopo_label") or ("Todos os colaboradores" if relatorio.get("escopo") == "todos" else "Equipe do gestor")),
        ("Matrícula de referência", relatorio.get("colaborador_referencia") or relatorio.get("gestor_referencia") or ""),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for row_idx, (label, value) in enumerate(metadados, start=3):
        ws_resumo.cell(row=row_idx, column=1, value=label).font = font_negrito
        ws_resumo.cell(row=row_idx, column=2, value=value)

    ws_resumo["A8"] = "Indicadores"
    ws_resumo["A8"].fill = fill_secao
    ws_resumo["A8"].font = font_negrito
    ws_resumo.merge_cells("A8:B8")

    indicadores = [
        ("Total de solicitações", "=MAX(COUNTA('Solicitações'!A:A)-1,0)"),
        ("Total de dias", "=SUM('Solicitações'!F:F)"),
        ("Colaboradores com lançamentos", int(relatorio.get("total_colaboradores_com_lancamento") or 0)),
        ("Aprovadas", '=COUNTIF(\'Solicitações\'!G:G,"*APROV*")'),
        ("Em análise", '=COUNTIF(\'Solicitações\'!G:G,"*ANÁLISE*")+COUNTIF(\'Solicitações\'!G:G,"*ANALISE*")'),
        ("Pendentes", '=COUNTIF(\'Solicitações\'!G:G,"PENDENTE")'),
        ("Reservadas", '=COUNTIF(\'Solicitações\'!G:G,"*RESERV*")'),
        ("Negadas/canceladas", '=COUNTIF(\'Solicitações\'!G:G,"*REPROV*")+COUNTIF(\'Solicitações\'!G:G,"*CANCEL*")+COUNTIF(\'Solicitações\'!G:G,"*NEGAD*")'),
    ]
    for row_idx, (label, value) in enumerate(indicadores, start=9):
        ws_resumo.cell(row=row_idx, column=1, value=label)
        ws_resumo.cell(row=row_idx, column=2, value=value)
        ws_resumo.cell(row=row_idx, column=1).border = borda_fina
        ws_resumo.cell(row=row_idx, column=2).border = borda_fina

    resumo_start = 19
    headers_resumo = ["Matrícula", "Colaborador", "Solicitações", "Total de dias", "Dias aprovados"]
    for col, header in enumerate(headers_resumo, start=1):
        cell = ws_resumo.cell(row=resumo_start, column=col, value=header)
        cell.fill = fill_cabecalho
        cell.font = font_branca
        cell.alignment = Alignment(horizontal="center")

    colaboradores = sorted(
        (relatorio.get("por_colaborador") or {}).values(),
        key=lambda item: (str(item.get("nome") or "").casefold(), str(item.get("matricula") or "")),
    )
    for row_idx, info in enumerate(colaboradores, start=resumo_start + 1):
        ws_resumo.cell(row=row_idx, column=1, value=info.get("matricula") or "")
        ws_resumo.cell(row=row_idx, column=2, value=info.get("nome") or "")
        ws_resumo.cell(row=row_idx, column=3, value=len(info.get("solicitacoes") or []))
        ws_resumo.cell(row=row_idx, column=4, value=float(info.get("total_dias") or 0))
        ws_resumo.cell(row=row_idx, column=5, value=float(info.get("total_aprovadas") or 0))
        ws_resumo.cell(row=row_idx, column=4).number_format = '0.00'
        ws_resumo.cell(row=row_idx, column=5).number_format = '0.00'

    headers = [
        "ID", "Matrícula", "Colaborador", "Data inicial", "Data final",
        "Dias", "Status", "Solicitação", "Tipo de saldo", "Observações",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws_detalhes.cell(row=1, column=col, value=header)
        cell.fill = fill_cabecalho
        cell.font = font_branca
        cell.alignment = Alignment(horizontal="center", vertical="center")

    detalhe_row = 2
    for info in colaboradores:
        for sol in info.get("solicitacoes") or []:
            valores = [
                sol.get("id"), info.get("matricula") or "", info.get("nome") or "",
                _parse_data_br(sol.get("inicio")), _parse_data_br(sol.get("fim")),
                float(sol.get("dias") or 0), sol.get("status") or "",
                sol.get("solicitacao") or "", sol.get("saldo_tipo") or "",
                sol.get("observacoes") or "",
            ]
            for col, value in enumerate(valores, start=1):
                ws_detalhes.cell(row=detalhe_row, column=col, value=value)
            ws_detalhes.cell(row=detalhe_row, column=4).number_format = "dd/mm/yyyy"
            ws_detalhes.cell(row=detalhe_row, column=5).number_format = "dd/mm/yyyy"
            ws_detalhes.cell(row=detalhe_row, column=6).number_format = "0.00"
            ws_detalhes.cell(row=detalhe_row, column=10).alignment = Alignment(wrap_text=True, vertical="top")
            detalhe_row += 1

    ws_resumo.freeze_panes = f"A{resumo_start + 1}"
    ws_resumo.auto_filter.ref = f"A{resumo_start}:E{max(resumo_start, resumo_start + len(colaboradores))}"
    ws_detalhes.freeze_panes = "A2"
    ws_detalhes.auto_filter.ref = f"A1:J{max(1, detalhe_row - 1)}"
    ws_detalhes.sheet_view.showGridLines = False
    ws_resumo.sheet_view.showGridLines = False
    ws_detalhes.row_dimensions[1].height = 26

    _auto_width(ws_resumo, max_width=42)
    _auto_width(ws_detalhes, max_width=55)
    ws_detalhes.column_dimensions["J"].width = 55

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
